# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import logging
from datetime import datetime, date, timedelta
import hashlib
import base64
import io

_logger = logging.getLogger(__name__)

class HrAnonymousMessage(models.Model):
    _name = 'hr.anonymous.message'
    _description = 'Anonymous HR Message'
    _order = 'create_date desc'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char(string='Subject', required=True)
    description = fields.Html(string='Message', required=True)

    category_id = fields.Many2one(
        'hr.anonymous.message.category',
        string='Category',
        required=True,
        ondelete='restrict',
        index=True,
    )
    
    category_legacy = fields.Selection([
        ('complaint', 'Complaint'),
        ('suggestion', 'Suggestion'),
        ('concern', 'Concern'),
        ('harassment', 'Harassment Report'),
        ('discrimination', 'Discrimination Report'),
        ('safety', 'Safety Issue'),
        ('ethics', 'Ethics Violation'),
        ('general', 'General Message'),
    ], string='Category (Legacy)', default='general', required=True)
    
    state = fields.Selection([
        ('draft', 'Draft'),
        ('sent', 'Sent'),
        ('acknowledged', 'Acknowledged'),
        ('in_progress', 'In Progress'),
        ('resolved', 'Resolved'),
        ('declined', 'Declined'),
        ('closed', 'Closed by Employee'),
    ], string='Status', default='draft', required=True, tracking=True)

    sender_audit_hash = fields.Char(
        string='Audit Hash', readonly=True,
        help='Encrypted audit trail - not viewable in UI'
    )
    sender_user_id = fields.Many2one(
        'res.users', string='Sender', readonly=True, copy=False,
    )
    priority = fields.Selection([
        ('0', 'Low'), ('1', 'Normal'), ('2', 'High'), ('3', 'Urgent'),
    ], string='Priority', default='1')
    
    hr_notes = fields.Text(
        string='HR Internal Notes', groups='hr.group_hr_user', tracking=True
    )
    resolution_notes = fields.Text(
        string='Resolution Notes',
        help='Details about how this message was resolved or why it was declined',
        groups='hr.group_hr_user', tracking=True
    )
    mail_sent = fields.Boolean(string='Email Sent', default=False, readonly=True)
    mail_id = fields.Many2one('mail.mail', string='Email Record', readonly=True)
    is_closed_by_employee = fields.Boolean(
        string='Closed by Employee', default=False, readonly=True
    )
    closed_date = fields.Datetime(string='Closed Date', readonly=True)

    date_period = fields.Selection([
        ('today',      'Today'),
        ('this_week',  'This Week'),
        ('this_month', 'This Month'),
        ('last_month', 'Last Month'),
        ('older',      'Older'),
    ], string='Date Period', compute='_compute_date_period', search='_search_date_period', store=True)

    is_my_message = fields.Boolean(
        string='My Message',
        compute='_compute_is_my_message',
        search='_search_is_my_message'
    )

    @api.depends('create_date')
    def _compute_date_period(self):
        today = date.today()
        # Start of this week (Monday)
        week_start = today - timedelta(days=today.weekday())
        # Start of this month
        month_start = today.replace(day=1)
        # Start of last month
        if today.month == 1:
            last_month_start = date(today.year - 1, 12, 1)
            last_month_end   = date(today.year, 1, 1)
        else:
            last_month_start = date(today.year, today.month - 1, 1)
            last_month_end   = month_start

        for rec in self:
            if not rec.create_date:
                rec.date_period = 'older'
                continue
            d = rec.create_date.date()
            if d == today:
                rec.date_period = 'today'
            elif d >= week_start:
                rec.date_period = 'this_week'
            elif d >= month_start:
                rec.date_period = 'this_month'
            elif d >= last_month_start:
                rec.date_period = 'last_month'
            else:
                rec.date_period = 'older'

    @api.model
    def _search_date_period(self, operator, value):
        today = date.today()
        week_start  = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)
        if today.month == 1:
            last_month_start = date(today.year - 1, 12, 1)
            last_month_end   = date(today.year, 1, 1)
        else:
            last_month_start = date(today.year, today.month - 1, 1)
            last_month_end   = month_start

        # Map each period key to a domain
        domains = {
            'today':      [('create_date', '>=', fields.Datetime.to_string(datetime.combine(today,            datetime.min.time()))),
                        ('create_date', '<',  fields.Datetime.to_string(datetime.combine(today + timedelta(days=1), datetime.min.time())))],
            'this_week':  [('create_date', '>=', fields.Datetime.to_string(datetime.combine(week_start,       datetime.min.time()))),
                        ('create_date', '<',  fields.Datetime.to_string(datetime.combine(today + timedelta(days=1), datetime.min.time())))],
            'this_month': [('create_date', '>=', fields.Datetime.to_string(datetime.combine(month_start,      datetime.min.time()))),
                        ('create_date', '<',  fields.Datetime.to_string(datetime.combine(today + timedelta(days=1), datetime.min.time())))],
            'last_month': [('create_date', '>=', fields.Datetime.to_string(datetime.combine(last_month_start, datetime.min.time()))),
                        ('create_date', '<',  fields.Datetime.to_string(datetime.combine(last_month_end,   datetime.min.time())))],
            'older':      [('create_date', '<',  fields.Datetime.to_string(datetime.combine(last_month_start, datetime.min.time())))],
        }

        if operator == '=' and value in domains:
            return domains[value]
        if operator == 'in':
            # combine with OR if multiple values selected
            from odoo.osv import expression
            return expression.OR([domains[v] for v in value if v in domains])
        return []

    @api.depends('sender_audit_hash')
    def _compute_is_my_message(self):
        for record in self:
            current_user_hash = self._generate_user_hash(self.env.user.id)
            record.is_my_message = record.sender_audit_hash == current_user_hash

    def _search_is_my_message(self, operator, value):
        current_user_hash = self._generate_user_hash(self.env.user.id)
        if operator == '=' and value:
            return [('sender_audit_hash', '=', current_user_hash)]
        return [('sender_audit_hash', '!=', current_user_hash)]

    def read(self, fields=None, load='_classic_read'):
        """Strip sender_user_id so it is never exposed via RPC"""
        if fields and 'sender_user_id' in fields:
            fields = [f for f in fields if f != 'sender_user_id']
        return super().read(fields=fields, load=load)

    @api.model
    def _generate_user_hash(self, user_id):
        secret = self.env['ir.config_parameter'].sudo().get_param(
            'database.secret', default='default_secret_change_in_production'
        )
        return hashlib.sha256(f"{user_id}_{secret}".encode()).hexdigest()

    def message_post(self, **kwargs):
        bot_partner = self.env.ref('base.partner_root')
        kwargs['author_id'] = bot_partner.id
        kwargs.pop('email_from', None)
        return super(HrAnonymousMessage, self.sudo()).message_post(**kwargs)

    def _message_compute_author(self, author_id=None, email_from=None, raise_on_email=True):
        bot_partner = self.env.ref('base.partner_root')
        company_email = self.env.company.email or 'noreply@localhost'
        return bot_partner.id, f"Anonymous HR System <{company_email}>"

    def _message_log(self, **kwargs):
        bot_partner = self.env.ref('base.partner_root')
        kwargs['author_id'] = bot_partner.id
        return super(HrAnonymousMessage, self.sudo())._message_log(**kwargs)

    def _message_notify_by_email(self, message, recipients_data, **kwargs):
        return super(HrAnonymousMessage, self.sudo())._message_notify_by_email(
            message, recipients_data, **kwargs
        )

    def send_to_hr(self):
        """
        Employee submits an anonymous message.

        What this does:
        - Saves the message with state = 'sent'
        - Creates an encrypted audit log entry
        - Notifies HR users inside Odoo (chatter + activity)
        - Does NOT send any email — emails are handled exclusively
          by the monthly cron job (_cron_send_monthly_report)
        """
        self.ensure_one()

        ICP = self.env['ir.config_parameter'].sudo()
        hr_email = ICP.get_param('hr_anonymous_message.hr_email', default='').strip()
        if not hr_email:
            raise UserError(_(
                'HR email is not configured. Please ask your administrator to '
                'configure it in Settings → HR Anonymous Messages.'
            ))

        _logger.info(
            f"Anonymous message '{self.name}' submitted — "
            f"Hash: {self.sender_audit_hash[:8]}... (no email sent, monthly cron will handle)"
        )

        self._create_audit_log()

        self.write({'state': 'sent', 'mail_sent': False})

        self._notify_hr_users()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Message Submitted!'),
                'message': _(
                    'Your anonymous message has been submitted to HR. '
                    'You will be notified when its status changes.'
                ),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }

    def _create_audit_log(self):
        self.env['hr.anonymous.message.audit'].sudo().create({
            'message_id': self.id,
            'user_hash': self.sender_audit_hash,
            'timestamp': fields.Datetime.now(),
            'action': 'message_sent',
        })

    def _notify_hr_users(self):
        """Notify HR users inside Odoo — no email, fully anonymous"""
        hr_group = self.env.ref('hr.group_hr_user', raise_if_not_found=False)
        if not hr_group:
            return
        hr_users = hr_group.user_ids
        if not hr_users:
            return

        partner_ids = hr_users.mapped('partner_id').ids

        self.sudo().message_post(
            body=_('A new anonymous message has been received: <strong>%s</strong>') % self.name,
            subject=_('New Anonymous Message'),
            partner_ids=partner_ids,
            message_type='notification',
            subtype_xmlid='mail.mt_comment',
        )
        self.sudo().activity_schedule(
            'mail.mail_activity_data_todo',
            summary=_('New Anonymous Message: %s') % self.name,
            note=_('Category: %s') % (self.category_id.name if self.category_id else ''),
            user_id=hr_users[0].id,
        )

    def _notify_employee_status_change(self, old_state):
        if old_state != self.state:
            status_label = dict(self._fields['state'].selection).get(self.state)
            self.message_post(
                body=_('Your message "%s" status changed to: %s') % (self.name, status_label),
                subject=_('Message Status Updated'),
                message_type='notification',
                subtype_xmlid='mail.mt_comment',
            )

    def action_acknowledge(self):
        self.ensure_one()
        old_state = self.state
        self.write({'state': 'acknowledged'})
        self._notify_employee_status_change(old_state)
        return True

    def action_in_progress(self):
        self.ensure_one()
        old_state = self.state
        self.write({'state': 'in_progress'})
        self._notify_employee_status_change(old_state)
        return True

    def action_resolve(self):
        self.ensure_one()
        old_state = self.state
        self.write({'state': 'resolved'})
        self._notify_employee_status_change(old_state)
        return True

    def action_decline(self):
        self.ensure_one()
        old_state = self.state
        self.write({'state': 'declined'})
        self._notify_employee_status_change(old_state)
        return True

    def action_close_ticket(self):
        self.ensure_one()
        if not self.is_my_message:
            raise UserError(_('You can only close your own messages.'))
        if self.state == 'draft':
            raise UserError(_('You cannot close a message that has not been sent yet.'))
        self.write({
            'state': 'closed',
            'is_closed_by_employee': True,
            'closed_date': fields.Datetime.now(),
        })
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Ticket Closed'),
                'message': _('Your message has been marked as closed.'),
                'type': 'success',
                'sticky': False,
            }
        }

    @api.model
    def create(self, vals_list):
        if isinstance(vals_list, dict):
            vals_list = [vals_list]
        real_user_id = self.env.user.id
        for vals in vals_list:
            vals['sender_audit_hash'] = self._generate_user_hash(real_user_id)
            vals['sender_user_id'] = real_user_id
        records = super(HrAnonymousMessage, self.sudo()).create(vals_list)
        for record in records:
            self.env.cr.execute(
                "UPDATE hr_anonymous_message SET sender_user_id = %s WHERE id = %s",
                (real_user_id, record.id)
            )
        self.env.cr.flush()
        records.invalidate_recordset(['sender_user_id'])
        return records

    def write(self, vals):
        for record in self:
            old_state = record.state
            res = super(HrAnonymousMessage, record).write(vals)
            if 'state' in vals and old_state != vals['state']:
                record._notify_employee_status_change(old_state)
            return res
        return super(HrAnonymousMessage, self).write(vals)

    @api.constrains('state')
    def _check_state_change_permission(self):
        for record in self:
            if not self.env.user.has_group('hr.group_hr_user') and \
               not self.env.user.has_group('base.group_system'):
                if record.state not in ['draft', 'sent', 'closed']:
                    raise ValidationError(_('Only HR users can change message status.'))

    def _generate_excel_export(self, messages):
        """
        Build a styled two-sheet Excel workbook for the monthly report.
        Sheet 1: full message list (no sender identity).
        Sheet 2: category + status summary.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError(_('openpyxl is required. Run: pip install openpyxl'))

        category_labels = {
            'complaint': 'Complaint', 'suggestion': 'Suggestion',
            'concern': 'Concern', 'harassment': 'Harassment Report',
            'discrimination': 'Discrimination Report', 'safety': 'Safety Issue',
            'ethics': 'Ethics Violation', 'general': 'General Message',
        }
        priority_labels = {'0': 'Low', '1': 'Normal', '2': 'High', '3': 'Urgent'}
        state_labels = {
            'draft': 'Draft', 'sent': 'Sent', 'acknowledged': 'Acknowledged',
            'in_progress': 'In Progress', 'resolved': 'Resolved',
            'declined': 'Declined', 'closed': 'Closed by Employee',
        }

        thin = Side(style='thin', color='CCCCCC')
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
        h_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        h_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
        h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        d_font = Font(name='Arial', size=10)
        d_align = Alignment(vertical='center', wrap_text=True)
        alt_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Anonymous Messages"

        headers = [
            'ID', 'Subject', 'Category', 'Priority', 'Status',
            'Date Submitted', 'Date Closed', 'Closed by Employee',
            'HR Notes', 'Resolution Notes',
        ]
        col_widths = [8, 35, 20, 12, 18, 20, 20, 18, 40, 40]

        for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=col, value=hdr)
            c.font = h_font
            c.fill = h_fill
            c.alignment = h_align
            c.border = bdr
            ws.column_dimensions[c.column_letter].width = w
        ws.row_dimensions[1].height = 30

        for row, msg in enumerate(messages, 2):
            row_data = [
                msg.id,
                msg.name or '',
                msg.category_id.name if msg.category_id else '',
                priority_labels.get(msg.priority, msg.priority or ''),
                state_labels.get(msg.state, msg.state or ''),
                msg.create_date.strftime('%Y-%m-%d %H:%M') if msg.create_date else '',
                msg.closed_date.strftime('%Y-%m-%d %H:%M') if msg.closed_date else '',
                'Yes' if msg.is_closed_by_employee else 'No',
                msg.hr_notes or '',
                msg.resolution_notes or '',
            ]
            use_alt = (row % 2 == 0)
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = d_font
                c.alignment = d_align
                c.border = bdr
                if use_alt:
                    c.fill = alt_fill
            ws.row_dimensions[row].height = 20

        ws2 = wb.create_sheet(title="Summary")
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 15

        ws2.merge_cells('A1:C1')
        ws2['A1'] = 'Anonymous Messages — Monthly Summary'
        ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='2E5090')
        ws2['A1'].alignment = Alignment(horizontal='center')
        ws2.row_dimensions[1].height = 30

        ws2.merge_cells('A2:C2')
        ws2['A2'] = f'Report Period: {date.today().strftime("%B %Y")}'
        ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
        ws2['A2'].alignment = Alignment(horizontal='center')

        # Category breakdown
        for col, hdr in enumerate(['Category', 'Count', 'Percentage'], 1):
            c = ws2.cell(row=4, column=col, value=hdr)
            c.font = Font(name='Arial', bold=True, color='FFFFFF')
            c.fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
            c.alignment = Alignment(horizontal='center')
            c.border = bdr

        cat_counts = {}
        for msg in messages:
            lbl = category_labels.get(msg.category, msg.category or 'Unknown')
            cat_counts[lbl] = cat_counts.get(lbl, 0) + 1

        total = len(messages)
        for i, (cat, cnt) in enumerate(cat_counts.items(), 5):
            ws2.cell(row=i, column=1, value=cat).border = bdr
            ws2.cell(row=i, column=2, value=cnt).border = bdr
            pct = ws2.cell(row=i, column=3,
                           value=f'=B{i}/B{5+len(cat_counts)}*100')
            pct.border = bdr
            pct.number_format = '0.0"%"'

        total_row = 5 + len(cat_counts)
        ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Arial', bold=True)
        ws2.cell(row=total_row, column=1).border = bdr
        ws2.cell(row=total_row, column=2,
                 value=f'=SUM(B5:B{total_row-1})').font = Font(name='Arial', bold=True)
        ws2.cell(row=total_row, column=2).border = bdr
        ws2.cell(row=total_row, column=3, value='100%').border = bdr

        ss = total_row + 3
        ws2.merge_cells(f'A{ss}:C{ss}')
        ws2.cell(row=ss, column=1,
                 value='Status Breakdown').font = Font(name='Arial', bold=True, size=11)

        for col, hdr in enumerate(['Status', 'Count', 'Percentage'], 1):
            c = ws2.cell(row=ss + 1, column=col, value=hdr)
            c.font = Font(name='Arial', bold=True, color='FFFFFF')
            c.fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
            c.alignment = Alignment(horizontal='center')
            c.border = bdr

        status_counts = {}
        for msg in messages:
            lbl = state_labels.get(msg.state, msg.state or 'Unknown')
            status_counts[lbl] = status_counts.get(lbl, 0) + 1

        for i, (st, cnt) in enumerate(status_counts.items(), ss + 2):
            ws2.cell(row=i, column=1, value=st).border = bdr
            ws2.cell(row=i, column=2, value=cnt).border = bdr
            ws2.cell(row=i, column=3,
                     value=f'{round(cnt/total*100,1)}%' if total else '0%').border = bdr

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @api.model
    def _cron_send_monthly_report(self):
        """
        Scheduled daily. On the configured day, sends a rich HTML
        statistics/analytics report to HR for the previous calendar month.
        No Excel attachment — the full dashboard is inside the email body.
        """
        _logger.info("=== Monthly Anonymous Messages Statistics Report: checking ===")

        ICP = self.env['ir.config_parameter'].sudo()

        enable_report = ICP.get_param(
            'hr_anonymous_message.enable_monthly_report', default='False'
        )
        if enable_report not in ('True', '1', 'true'):
            _logger.info("Monthly reports disabled in settings. Skipping.")
            return

        report_day = int(ICP.get_param(
            'hr_anonymous_message.monthly_report_day', default='1'
        ))
        today = date.today()
        if today.day != report_day:
            _logger.info(
                f"Today is day {today.day}, report day is {report_day}. Skipping."
            )
            return

        hr_email = ICP.get_param('hr_anonymous_message.hr_email', default='').strip()
        if not hr_email:
            _logger.error("Monthly report: HR email not configured. Aborting.")
            return

        # ── Determine previous month range ────────────────────────────────────
        if today.month == 1:
            report_month, report_year = 12, today.year - 1
        else:
            report_month, report_year = today.month - 1, today.year

        month_start = date(report_year, report_month, 1)
        month_end = date(
            report_year + 1 if report_month == 12 else report_year,
            1 if report_month == 12 else report_month + 1,
            1,
        )
        month_name = month_start.strftime('%B %Y')

        messages = self.sudo().search([
            ('create_date', '>=', fields.Datetime.to_string(
                datetime.combine(month_start, datetime.min.time())
            )),
            ('create_date', '<', fields.Datetime.to_string(
                datetime.combine(month_end, datetime.min.time())
            )),
            ('state', '!=', 'draft'),
        ])

        _logger.info(f"Building statistics report: {len(messages)} messages for {month_name}")

        email_body = self._build_statistics_email(messages, month_name, today)

        try:
            mail = self.env['mail.mail'].sudo().create({
                'subject': f'📊 HR Anonymous Messages — Monthly Report: {month_name}',
                'email_to': hr_email,
                'email_from': self.env.company.email or hr_email,
                'body_html': email_body,
                'auto_delete': True,
            })
            mail.sudo().send()
            _logger.info(f"Monthly statistics report sent to {hr_email} for {month_name}")
        except Exception as e:
            _logger.error(f"Failed to send monthly statistics report: {e}")
            _logger.exception("Traceback:")

    @api.model
    def _build_statistics_email(self, messages, month_name, today, is_test=False):
        """
        Build and return the full HTML statistics report email body.
        Called by both the monthly cron and the Settings test button.
        """
        from collections import Counter

        total = len(messages)

        # ── Status counts ─────────────────────────────────────────────────────
        state_counts = Counter(m.state for m in messages)
        resolved    = state_counts.get('resolved', 0)
        in_progress = state_counts.get('in_progress', 0)
        acknowledged= state_counts.get('acknowledged', 0)
        sent        = state_counts.get('sent', 0)
        declined    = state_counts.get('declined', 0)
        closed      = state_counts.get('closed', 0)

        open_count  = sent + acknowledged + in_progress
        closed_total= resolved + declined + closed

        # Resolution rate (resolved out of all completed)
        if closed_total > 0:
            resolution_rate = round((resolved / closed_total) * 100)
        else:
            resolution_rate = 0

        # Completion rate (anything not still open)
        completion_rate = round((closed_total / total) * 100) if total > 0 else 0

        # ── Category breakdown ────────────────────────────────────────────────
        cat_counter = Counter()
        for m in messages:
            label = m.category_id.name if m.category_id else 'Uncategorised'
            cat_counter[label] += 1
        top_categories = cat_counter.most_common()

        # ── Priority breakdown ────────────────────────────────────────────────
        priority_labels = {'0': 'Low', '1': 'Normal', '2': 'High', '3': 'Urgent'}
        priority_counter = Counter()
        for m in messages:
            priority_counter[priority_labels.get(m.priority, 'Normal')] += 1

        # ── Trending / similar clusters ───────────────────────────────────────
        # Group messages by category — categories with 3+ messages are "trending"
        trending = [
            (cat, count) for cat, count in top_categories if count >= 3
        ]

        # ── Messages still open going into next month ─────────────────────────
        still_open = messages.filtered(
            lambda m: m.state in ['sent', 'acknowledged', 'in_progress']
        )

        # ── Urgent messages count ─────────────────────────────────────────────
        urgent_count = len(messages.filtered(lambda m: m.priority == '3'))
        high_count   = len(messages.filtered(lambda m: m.priority == '2'))

        # ── Helper: percentage bar HTML ───────────────────────────────────────
        def pct(count):
            return round((count / total) * 100) if total > 0 else 0

        def bar(count, color='#2E5090'):
            width = pct(count)
            return (
                f'<div style="background:#e9ecef;border-radius:4px;height:10px;'
                f'margin-top:4px;">'
                f'<div style="background:{color};width:{width}%;height:10px;'
                f'border-radius:4px;"></div></div>'
            )

        # ── Stat card helper ──────────────────────────────────────────────────
        def stat_card(emoji, label, value, color='#2E5090', sub=''):
            return f'''
            <td style="width:25%;padding:8px;">
              <div style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;
                          padding:16px;text-align:center;border-top:4px solid {color};">
                <div style="font-size:24px;">{emoji}</div>
                <div style="font-size:28px;font-weight:bold;color:{color};
                            margin:6px 0;">{value}</div>
                <div style="font-size:12px;color:#666;">{label}</div>
                {f'<div style="font-size:11px;color:#999;margin-top:4px;">{sub}</div>' if sub else ''}
              </div>
            </td>'''

        # ── Category rows ─────────────────────────────────────────────────────
        category_colors = [
            '#2E5090','#4A90D9','#27AE60','#E67E22',
            '#8E44AD','#E74C3C','#1ABC9C','#F39C12',
        ]
        cat_rows = ''
        for i, (cat, count) in enumerate(top_categories):
            color = category_colors[i % len(category_colors)]
            cat_rows += f'''
            <tr>
              <td style="padding:10px 0;font-size:13px;color:#333;width:35%;">
                <span style="display:inline-block;width:10px;height:10px;
                             background:{color};border-radius:50%;
                             margin-right:6px;"></span>{cat}
              </td>
              <td style="padding:10px 0;width:50%;">
                {bar(count, color)}
              </td>
              <td style="padding:10px 0;text-align:right;font-size:13px;
                         font-weight:bold;color:{color};width:15%;">
                {count} <span style="font-weight:normal;color:#999;
                               font-size:11px;">({pct(count)}%)</span>
              </td>
            </tr>'''

        # ── Priority rows ─────────────────────────────────────────────────────
        priority_colors = {
            'Urgent': '#C0392B', 'High': '#E67E22',
            'Normal': '#2E5090', 'Low':  '#27AE60',
        }
        priority_rows = ''
        for label in ['Urgent', 'High', 'Normal', 'Low']:
            count = priority_counter.get(label, 0)
            color = priority_colors[label]
            priority_rows += f'''
            <tr>
              <td style="padding:8px 0;font-size:13px;color:#333;width:20%;">
                {label}
              </td>
              <td style="padding:8px 0;width:60%;">
                {bar(count, color)}
              </td>
              <td style="padding:8px 0;text-align:right;font-size:13px;
                         font-weight:bold;color:{color};width:20%;">
                {count}
              </td>
            </tr>'''

        # ── Trending issues block ─────────────────────────────────────────────
        if trending:
            trending_items = ''.join([
                f'''<div style="display:inline-block;background:#fff3cd;
                               border:1px solid #ffc107;border-radius:20px;
                               padding:4px 12px;margin:4px;font-size:12px;
                               color:#856404;">
                    🔥 {cat} <strong>({count})</strong>
                  </div>'''
                for cat, count in trending
            ])
            trending_block = f'''
            <div style="background:#fffbf0;border:1px solid #ffc107;
                        border-radius:8px;padding:20px;margin:20px 0;">
              <h3 style="margin:0 0 12px;color:#856404;font-size:15px;">
                🔥 Trending Issues This Month
              </h3>
              <p style="font-size:12px;color:#999;margin:0 0 10px;">
                Categories with 3 or more messages — these may need systemic attention
              </p>
              {trending_items}
            </div>'''
        else:
            trending_block = ''

        # ── Still open block ──────────────────────────────────────────────────
        if still_open:
            open_by_cat = Counter(
                m.category_id.name if m.category_id else 'Uncategorised'
                for m in still_open
            )
            open_rows = ''.join([
                f'<li style="font-size:13px;color:#555;padding:2px 0;">'
                f'{cat}: <strong>{cnt}</strong></li>'
                for cat, cnt in open_by_cat.most_common()
            ])
            open_block = f'''
            <div style="background:#fdecea;border:1px solid #f5c6cb;
                        border-radius:8px;padding:20px;margin:20px 0;">
              <h3 style="margin:0 0 10px;color:#c0392b;font-size:15px;">
                ⚠️ {len(still_open)} Message(s) Still Open — Carried into Next Month
              </h3>
              <p style="font-size:12px;color:#999;margin:0 0 10px;">
                These messages have not yet been resolved or closed
              </p>
              <ul style="margin:0;padding-left:16px;">
                {open_rows}
              </ul>
            </div>'''
        else:
            open_block = f'''
            <div style="background:#e8f5e9;border:1px solid #a5d6a7;
                        border-radius:8px;padding:16px;margin:20px 0;">
              <p style="margin:0;font-size:14px;color:#2e7d32;">
                ✅ All messages from this month have been resolved or closed.
                Excellent work!
              </p>
            </div>'''

        # ── Urgent alert block ────────────────────────────────────────────────
        urgent_block = ''
        if urgent_count > 0:
            urgent_block = f'''
            <div style="background:#fdecea;border-left:4px solid #C0392B;
                        border-radius:4px;padding:14px 16px;margin:16px 0;">
              <p style="margin:0;font-size:13px;color:#c0392b;">
                🚨 <strong>{urgent_count} Urgent</strong> and
                <strong>{high_count} High Priority</strong> message(s) were
                received this month. Please ensure these were handled promptly.
              </p>
            </div>'''

        # ── Test banner ───────────────────────────────────────────────────────
        test_banner = ''
        if is_test:
            test_banner = '''
            <div style="background:#fff3e0;border:2px solid #ff8f00;
                        border-radius:6px;padding:12px 16px;margin-bottom:20px;">
              <p style="margin:0;font-size:13px;color:#e65100;">
                🧪 <strong>TEST REPORT</strong> — This was sent manually from
                Settings and covers the <strong>current month</strong>.
                The scheduled cron sends the previous month automatically.
              </p>
            </div>'''

        # ── Status breakdown rows ─────────────────────────────────────────────
        status_items = [
            ('📨', 'Sent / Awaiting Review', sent,        '#E67E22'),
            ('👀', 'Acknowledged',            acknowledged,'#3498DB'),
            ('🔄', 'In Progress',             in_progress, '#8E44AD'),
            ('✅', 'Resolved',                resolved,    '#27AE60'),
            ('❌', 'Declined',                declined,    '#E74C3C'),
            ('🔒', 'Closed by Employee',      closed,      '#7F8C8D'),
        ]
        status_rows = ''
        for emoji, label, count, color in status_items:
            status_rows += f'''
            <tr style="border-bottom:1px solid #f0f0f0;">
              <td style="padding:10px 0;font-size:13px;color:#333;">
                {emoji} {label}
              </td>
              <td style="padding:10px 0;width:50%;">
                {bar(count, color)}
              </td>
              <td style="padding:10px 0;text-align:right;font-size:14px;
                         font-weight:bold;color:{color};">
                {count}
              </td>
            </tr>'''

        # ── Assemble the full email ────────────────────────────────────────────
        html = f"""
<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f0f2f5;font-family:Arial,sans-serif;">
<div style="max-width:680px;margin:0 auto;padding:24px 16px;">

  <!-- Header -->
  <div style="background:linear-gradient(135deg,#1a3a6b 0%,#2E5090 60%,#4A90D9 100%);
              padding:36px 32px;border-radius:12px 12px 0 0;text-align:center;">
    <div style="font-size:42px;margin-bottom:10px;">📊</div>
    <h1 style="color:#fff;margin:0;font-size:24px;font-weight:bold;
               letter-spacing:-0.5px;">
      Anonymous HR Messages
    </h1>
    <p style="color:#a8c8f0;margin:8px 0 0;font-size:14px;">
      Monthly Statistics Report &nbsp;·&nbsp; {month_name}
    </p>
    <p style="color:#7aabdf;margin:6px 0 0;font-size:11px;">
      Generated on {today.strftime('%d %B %Y')} &nbsp;·&nbsp; Confidential
    </p>
  </div>

  <!-- Body -->
  <div style="background:#fff;padding:32px;
              border-left:1px solid #dde3ec;border-right:1px solid #dde3ec;">

    {test_banner}

    <p style="font-size:15px;color:#333;margin-top:0;">Dear HR Team,</p>
    <p style="font-size:14px;color:#666;line-height:1.7;margin-bottom:24px;">
      Here is your anonymous messaging statistics summary for
      <strong>{month_name}</strong>. All sender identities remain protected
      and are not included in this report.
    </p>

    {urgent_block}

    <!-- Top-level stat cards -->
    <table style="width:100%;border-collapse:collapse;margin-bottom:8px;">
      <tr>
        {stat_card('📬', 'Total Messages',    total,           '#2E5090')}
        {stat_card('✅', 'Resolved',           resolved,        '#27AE60',
                   f'{resolution_rate}% resolution rate')}
        {stat_card('⏳', 'Still Open',         open_count,      '#E67E22',
                   'carried to next month' if open_count else 'all clear!')}
        {stat_card('🔒', 'Closed by Employee', closed,          '#7F8C8D')}
      </tr>
    </table>

    <!-- Completion rate pill -->
    <div style="text-align:center;margin:16px 0 28px;">
      <span style="background:#e8f5e9;border:1px solid #a5d6a7;border-radius:20px;
                   padding:6px 16px;font-size:13px;color:#2e7d32;">
        📈 Overall completion rate this month: <strong>{completion_rate}%</strong>
      </span>
    </div>

    <!-- Status Breakdown -->
    <div style="background:#f8f9fa;border:1px solid #e9ecef;border-radius:8px;
                padding:20px;margin-bottom:20px;">
      <h3 style="margin:0 0 16px;color:#1a3a6b;font-size:15px;
                 border-bottom:2px solid #2E5090;padding-bottom:8px;">
        📋 Status Breakdown
      </h3>
      <table style="width:100%;border-collapse:collapse;">
        {status_rows}
      </table>
    </div>

    <!-- Category Breakdown -->
    <div style="background:#f8f9fa;border:1px solid #e9ecef;border-radius:8px;
                padding:20px;margin-bottom:20px;">
      <h3 style="margin:0 0 6px;color:#1a3a6b;font-size:15px;
                 border-bottom:2px solid #2E5090;padding-bottom:8px;">
        🏷️ Messages by Category
      </h3>
      <p style="font-size:12px;color:#999;margin:0 0 14px;">
        Which topics employees raised most this month
      </p>
      <table style="width:100%;border-collapse:collapse;">
        {cat_rows if cat_rows else '<tr><td style="color:#999;font-size:13px;padding:8px 0;">No categorised messages this month.</td></tr>'}
      </table>
    </div>

    <!-- Priority Breakdown -->
    <div style="background:#f8f9fa;border:1px solid #e9ecef;border-radius:8px;
                padding:20px;margin-bottom:20px;">
      <h3 style="margin:0 0 6px;color:#1a3a6b;font-size:15px;
                 border-bottom:2px solid #2E5090;padding-bottom:8px;">
        🎯 Messages by Priority
      </h3>
      <p style="font-size:12px;color:#999;margin:0 0 14px;">
        How employees rated the urgency of their submissions
      </p>
      <table style="width:100%;border-collapse:collapse;">
        {priority_rows}
      </table>
    </div>

    <!-- Trending Issues -->
    {trending_block}

    <!-- Still Open / All Clear -->
    {open_block}

    <!-- Key Takeaways -->
    <div style="background:#f0f4ff;border:1px solid #c5d3f0;border-radius:8px;
                padding:20px;margin-bottom:20px;">
      <h3 style="margin:0 0 12px;color:#1a3a6b;font-size:15px;">
        💡 Key Takeaways
      </h3>
      <ul style="margin:0;padding-left:18px;font-size:13px;color:#444;
                 line-height:1.9;">
        <li>
          <strong>{total}</strong> anonymous message(s) were submitted in {month_name}
        </li>
        <li>
          <strong>{resolved}</strong> were fully resolved
          ({resolution_rate}% of completed cases)
        </li>
        {'<li>The most active category was <strong>' + top_categories[0][0] + '</strong> with <strong>' + str(top_categories[0][1]) + '</strong> message(s)</li>' if top_categories else ''}
        {'<li style="color:#c0392b;"><strong>' + str(urgent_count) + ' urgent</strong> message(s) required immediate attention</li>' if urgent_count else '<li>No urgent messages this month 🎉</li>'}
        <li>
          <strong>{open_count}</strong> message(s) remain open and will carry
          into the next reporting period
        </li>
      </ul>
    </div>

    <!-- Privacy Notice -->
    <div style="background:#fff3cd;border:1px solid #ffc107;border-radius:6px;
                padding:12px 16px;">
      <p style="margin:0;font-size:12px;color:#856404;">
        🛡️ <strong>Privacy Notice:</strong> This report contains no sender
        identity information. All submissions remain fully anonymous in
        compliance with company policy. Message content is not included.
      </p>
    </div>

  </div>

  <!-- Footer -->
  <div style="background:#1a3a6b;padding:18px 32px;text-align:center;
              border-radius:0 0 12px 12px;">
    <p style="margin:0;font-size:12px;color:#a8c8f0;">
      HR Anonymous Messaging System &nbsp;·&nbsp; Automated Monthly Report
    </p>
    <p style="margin:4px 0 0;font-size:11px;color:#6a96c8;">
      Do not reply to this email &nbsp;·&nbsp; Confidential HR Document
    </p>
  </div>

</div>
</body>
</html>"""

        return html

class HrAnonymousMessageAudit(models.Model):
    """Separate audit log table — NOT accessible via UI"""
    _name = 'hr.anonymous.message.audit'
    _description = 'Anonymous Message Audit Log (Encrypted)'
    _rec_name = 'message_id'

    message_id = fields.Many2one(
        'hr.anonymous.message', string='Message',
        required=True, ondelete='cascade'
    )
    user_hash = fields.Char(
        string='User Hash', required=True,
        help='Encrypted — cannot be reversed without system secret'
    )
    timestamp = fields.Datetime(string='Timestamp', required=True)
    action = fields.Char(string='Action', required=True)